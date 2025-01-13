from openpyxl import load_workbook
from openpyxl.styles import Alignment
import glob

import ExcelVisualConverterConstantsVariables


def excel_visual_converter(files, columnWidth):
    for file in files:
        workBook = load_workbook(file)
        workSheet = workBook.active

        for col in workSheet.iter_cols():
            col_letter = col[0].column_letter
            workSheet.column_dimensions[col_letter].width = columnWidth

        for row in workSheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal=ExcelVisualConverterConstantsVariables.horizontalAlignment,
                                           vertical=ExcelVisualConverterConstantsVariables.verticalAlignment,
                                           wrap_text=ExcelVisualConverterConstantsVariables.wrapText)

        for row in range(1, workSheet.max_row + 1):
            workSheet.row_dimensions[row].height = ExcelVisualConverterConstantsVariables.rowHeight

        workBook.save(file)
        print(f"Cells parameters in {file} file converted, columns width are {columnWidth} ")


def excel_file_finder(whichFile):
    files = []

    if whichFile == "all":
        files = glob.glob("*.xlsx")
    elif isinstance(whichFile, list):
        for file_pattern in whichFile:
            files.extend(glob.glob(f"{file_pattern}"))
    else:
        files = glob.glob(f"{whichFile}")

    if files:
        return files
    else:
        print(f'No *.xlsx file in directory')


def main(workingMode="normal"):
    modes = ExcelVisualConverterConstantsVariables.mode_mapping.get(workingMode, [])

    for file_pattern, width in modes:
        xlsxFiles = excel_file_finder(file_pattern)
        excel_visual_converter(xlsxFiles, ExcelVisualConverterConstantsVariables.columnsWidth[width])


if __name__ == "__main__":
    main()
